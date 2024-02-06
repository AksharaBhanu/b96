import openpyxl

wb=openpyxl.open("../data/Book1.xlsx")
sheet=wb['Sheet2']
rc=sheet.max_row
cc=sheet.max_column

for i in range(1,rc+1):
    for j in range(1,cc+1):
        v=sheet.cell(i,j).value
        print(v,end=' ')
    print()

wb.close()