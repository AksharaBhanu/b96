import openpyxl
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By

path='../data/Orange.xlsx'
wb=openpyxl.open(path)
sheet=wb['Login']
rc=sheet.max_row
for i in range(2,rc+1):
    un=sheet.cell(i,1).value
    pw=sheet.cell(i,2).value
    expected=sheet.cell(i,3).value

    driver = Chrome()
    driver.implicitly_wait(10)
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

    driver.find_element(By.NAME,"username").send_keys(un)
    driver.find_element(By.NAME,"password").send_keys(pw)
    driver.find_element(By.XPATH,"//button[@type='submit']").click()

    actual=driver.current_url
    sheet.cell(i,4).value=actual

    if expected in actual:
        sheet.cell(i, 5).value = 'PASS'
        sheet.cell(i, 6).value = 'Home Page is Displayed'
    else:
        sheet.cell(i, 5).value = 'FAIL'
        sheet.cell(i, 6).value = 'Home Page is NOT Displayed'
    driver.quit()

wb.save(path)
wb.close()
