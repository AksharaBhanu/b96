import openpyxl
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By

wb=openpyxl.open('../data/HRMS.xlsx')
sheet=wb['Login']
un=sheet.cell(2,1).value
pw=sheet.cell(2,2).value
expected=sheet.cell(2,3).value

driver = Chrome()
driver.implicitly_wait(10)
driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

driver.find_element(By.NAME,"username").send_keys(un)
driver.find_element(By.NAME,"password").send_keys(pw)
driver.find_element(By.XPATH,"//button[@type='submit']").click()

actual=driver.current_url
sheet.cell(2,4).value=actual

if expected in actual:
    sheet.cell(2, 5).value = 'PASS'
    sheet.cell(2, 6).value = 'Home Page is Displayed'
else:
    sheet.cell(2, 5).value = 'FAIL'
    sheet.cell(2, 6).value = 'Home Page is NOT Displayed'

wb.save('../data/HRMS.xlsx')
wb.close()
